"""Export router — PDF and Excel generation.
See API_CONTRACTS.md for full specification.

Uses openpyxl for Excel, reportlab for PDF.
Falls back to CSV/text if libraries not available.
"""

from uuid import UUID
from datetime import date, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func as sa_func

from api.database import get_db
from api.roles import require_management, require_owner
from api.models import (
    ProductionOrder, OrderPosition, ProductionOrderItem,
    FinancialEntry, OrderFinancial,
    Material, MaterialStock, QualityCheck, User,
)

router = APIRouter()


@router.get("/materials/excel")
async def export_materials_excel(
    factory_id: UUID | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_management),
):
    """Export materials data to Excel (XLSX)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    # Query materials with stock info for the given factory
    query = db.query(Material).order_by(Material.name)
    materials = query.limit(1000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Materials"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Name", "Type", "Unit", "Current Balance",
        "Min Balance", "Supplier",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row, mat in enumerate(materials, 2):
        # Get stock for the specified factory (or first available)
        stock = None
        if factory_id:
            stock = (
                db.query(MaterialStock)
                .filter(MaterialStock.material_id == mat.id, MaterialStock.factory_id == factory_id)
                .first()
            )
        else:
            stock = (
                db.query(MaterialStock)
                .filter(MaterialStock.material_id == mat.id)
                .first()
            )

        ws.cell(row=row, column=1, value=mat.name).border = thin_border
        ws.cell(row=row, column=2, value=mat.material_type).border = thin_border
        ws.cell(row=row, column=3, value=mat.unit).border = thin_border
        ws.cell(row=row, column=4, value=float(stock.balance) if stock and stock.balance else 0).border = thin_border
        ws.cell(row=row, column=5, value=float(stock.min_balance) if stock and stock.min_balance else 0).border = thin_border
        ws.cell(row=row, column=6, value=mat.supplier.name if mat.supplier else "").border = thin_border

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=materials_{date.today()}.xlsx"},
    )


@router.get("/quality/excel")
async def export_quality_excel(
    factory_id: UUID | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_management),
):
    """Export quality inspection data to Excel (XLSX)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    # Default date range: last 30 days
    dt_to = date.fromisoformat(date_to) if date_to else date.today()
    dt_from = date.fromisoformat(date_from) if date_from else dt_to - timedelta(days=30)

    query = db.query(QualityCheck).filter(
        sa_func.date(QualityCheck.created_at) >= dt_from,
        sa_func.date(QualityCheck.created_at) <= dt_to,
    )
    if factory_id:
        query = query.filter(QualityCheck.factory_id == factory_id)

    inspections = query.order_by(QualityCheck.created_at.desc()).limit(1000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Quality Inspections"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Position", "Inspector", "Stage", "Result",
        "Defect Cause", "Notes", "Date",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row, qc in enumerate(inspections, 2):
        # Position info
        pos_label = ""
        if qc.position:
            order = qc.position.order if hasattr(qc.position, 'order') else None
            order_num = order.order_number if order else ""
            pos_label = f"{order_num} / {qc.position.color} / {qc.position.size}"

        # Inspector name
        inspector_name = ""
        if qc.checked_by:
            user = db.query(User).filter(User.id == qc.checked_by).first()
            inspector_name = user.name if user else ""

        result_val = qc.result if isinstance(qc.result, str) else qc.result.value
        stage_val = qc.stage if isinstance(qc.stage, str) else qc.stage.value

        # Defect cause
        defect_info = ""
        if qc.defect_cause:
            defect_info = qc.defect_cause.description or qc.defect_cause.code

        ws.cell(row=row, column=1, value=pos_label).border = thin_border
        ws.cell(row=row, column=2, value=inspector_name).border = thin_border
        ws.cell(row=row, column=3, value=stage_val).border = thin_border
        ws.cell(row=row, column=4, value=result_val).border = thin_border
        ws.cell(row=row, column=5, value=defect_info).border = thin_border
        ws.cell(row=row, column=6, value=qc.notes or "").border = thin_border
        ws.cell(row=row, column=7, value=str(qc.created_at.date()) if qc.created_at else "").border = thin_border

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=quality_{date.today()}.xlsx"},
    )


def _get_orders(db: Session, factory_id: UUID | None = None) -> list:
    """Get orders for export."""
    q = db.query(ProductionOrder).order_by(ProductionOrder.created_at.desc())
    if factory_id:
        q = q.filter(ProductionOrder.factory_id == factory_id)
    return q.limit(500).all()


@router.get("/orders/excel")
async def export_orders_excel(
    factory_id: UUID | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_management),
):
    """Export orders to Excel (XLSX)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    orders = _get_orders(db, factory_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "Order #", "Client", "Factory", "Status",
        "Created", "Deadline", "Schedule Deadline",
        "Shipped", "Source",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.order_number).border = thin_border
        ws.cell(row=row, column=2, value=order.client).border = thin_border
        ws.cell(row=row, column=3, value=order.factory.name if order.factory else "").border = thin_border
        status_val = order.status if isinstance(order.status, str) else order.status.value
        ws.cell(row=row, column=4, value=status_val).border = thin_border
        ws.cell(row=row, column=5, value=str(order.created_at.date()) if order.created_at else "").border = thin_border
        ws.cell(row=row, column=6, value=str(order.final_deadline) if order.final_deadline else "").border = thin_border
        ws.cell(row=row, column=7, value=str(order.schedule_deadline) if order.schedule_deadline else "").border = thin_border
        ws.cell(row=row, column=8, value=str(order.shipped_at.date()) if order.shipped_at else "").border = thin_border
        source_val = order.source if isinstance(order.source, str) else order.source.value
        ws.cell(row=row, column=9, value=source_val).border = thin_border

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=orders_{date.today()}.xlsx"},
    )


@router.get("/orders/pdf")
async def export_orders_pdf(
    factory_id: UUID | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_management),
):
    """Export orders to PDF."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise HTTPException(500, "reportlab not installed. Run: pip install reportlab")

    orders = _get_orders(db, factory_id)

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(f"Orders Report — {date.today()}", styles["Title"]))
    elements.append(Spacer(1, 12))

    # Table data
    data = [["Order #", "Client", "Status", "Deadline", "Shipped"]]
    for order in orders:
        status_val = order.status if isinstance(order.status, str) else order.status.value
        data.append([
            order.order_number,
            (order.client or "")[:30],
            status_val,
            str(order.final_deadline) if order.final_deadline else "",
            str(order.shipped_at.date()) if order.shipped_at else "",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B5797")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=orders_{date.today()}.pdf"},
    )


@router.get("/positions/pdf")
async def export_positions_pdf(
    order_id: UUID | None = None,
    factory_id: UUID | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_management),
):
    """Export positions to PDF."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        raise HTTPException(500, "reportlab not installed. Run: pip install reportlab")

    q = db.query(OrderPosition).join(ProductionOrder)
    if order_id:
        q = q.filter(OrderPosition.order_id == order_id)
    if factory_id:
        q = q.filter(OrderPosition.factory_id == factory_id)
    positions = q.order_by(OrderPosition.created_at.desc()).limit(500).all()

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    title = f"Positions Report — {date.today()}"
    if order_id:
        order = db.query(ProductionOrder).get(order_id)
        if order:
            title = f"Positions for Order {order.order_number}"

    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Spacer(1, 12))

    data = [["Order #", "Color", "Size", "Qty", "SQM", "Status", "Batch"]]
    for p in positions:
        status_val = p.status if isinstance(p.status, str) else p.status.value
        data.append([
            p.order.order_number if p.order else "",
            p.color,
            p.size,
            str(p.quantity),
            f"{float(p.quantity_sqm or 0):.2f}",
            status_val,
            str(p.batch_id)[:8] if p.batch_id else "",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B5797")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=positions_{date.today()}.pdf"},
    )


@router.post("/owner-monthly")
async def owner_monthly_report(
    factory_id: UUID | None = None,
    month: str | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_owner),
):
    """Owner monthly report with KPIs + financial summary."""
    from business.services.daily_kpi import calculate_dashboard_summary

    if month:
        year, m = int(month[:4]), int(month[5:7])
        d_from = date(year, m, 1)
        if m == 12:
            d_to = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            d_to = date(year, m + 1, 1) - timedelta(days=1)
    else:
        d_to = date.today()
        d_from = d_to.replace(day=1)

    summary = calculate_dashboard_summary(db, factory_id, d_from, d_to)

    # Financial breakdown
    financial_q = db.query(
        FinancialEntry.entry_type,
        FinancialEntry.category,
        sa_func.sum(FinancialEntry.amount).label("total"),
    ).filter(
        FinancialEntry.entry_date >= d_from,
        FinancialEntry.entry_date <= d_to,
    )
    if factory_id:
        financial_q = financial_q.filter(FinancialEntry.factory_id == factory_id)
    financial_q = financial_q.group_by(FinancialEntry.entry_type, FinancialEntry.category)

    financials = {}
    for row in financial_q.all():
        entry_type = row.entry_type if isinstance(row.entry_type, str) else row.entry_type.value
        category = row.category if isinstance(row.category, str) else row.category.value
        if entry_type not in financials:
            financials[entry_type] = {}
        financials[entry_type][category] = float(row.total or 0)

    # Revenue from shipped orders
    revenue_q = db.query(
        sa_func.sum(OrderFinancial.total_price)
    ).join(ProductionOrder).filter(
        ProductionOrder.shipped_at.isnot(None),
        sa_func.date(ProductionOrder.shipped_at) >= d_from,
        sa_func.date(ProductionOrder.shipped_at) <= d_to,
    )
    if factory_id:
        revenue_q = revenue_q.filter(ProductionOrder.factory_id == factory_id)
    revenue = float(revenue_q.scalar() or 0)

    return {
        "period": {"from": str(d_from), "to": str(d_to)},
        "kpi": summary,
        "financial_breakdown": financials,
        "revenue": revenue,
    }


@router.post("/ceo-daily")
async def ceo_daily_report(
    factory_id: UUID | None = None,
    report_date: str | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_management),
):
    """CEO daily summary report."""
    from business.services.daily_kpi import (
        calculate_dashboard_summary,
        calculate_production_metrics,
        calculate_material_metrics,
        get_activity_feed,
    )

    d = date.fromisoformat(report_date) if report_date else date.today()
    d_from = d
    d_to = d

    summary = calculate_dashboard_summary(db, factory_id, d_from, d_to)
    production = calculate_production_metrics(db, factory_id, d_from, d_to)
    materials = calculate_material_metrics(db, factory_id)
    activity = get_activity_feed(db, factory_id, limit=20)

    return {
        "date": str(d),
        "kpi": summary,
        "production": production,
        "material_deficits": materials,
        "recent_activity": activity,
    }


# ---------------------------------------------------------------------------
# Shared Excel styling helpers
# ---------------------------------------------------------------------------

def _xl_styles():
    """Return reusable openpyxl style objects."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    return {
        "header_font": Font(bold=True, color="FFFFFF", size=11),
        "header_fill": PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid"),
        "section_font": Font(bold=True, color="2B5797", size=13),
        "kpi_label_font": Font(bold=True, size=10),
        "kpi_value_font": Font(bold=True, size=12, color="2B5797"),
        "thin_border": Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        ),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
        "alt_fill": PatternFill(start_color="F2F6FC", end_color="F2F6FC", fill_type="solid"),
    }


def _write_header_row(ws, row: int, headers: list[str], styles: dict):
    """Write a styled header row and return row+1."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["center"]
        cell.border = styles["thin_border"]
    return row + 1


def _write_data_row(ws, row: int, values: list, styles: dict, alt: bool = False):
    """Write one data row with styling."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = styles["thin_border"]
        cell.alignment = styles["left"]
        if alt:
            cell.fill = styles["alt_fill"]
    return row + 1


def _auto_width(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 3, 45)


# ---------------------------------------------------------------------------
# CEO Daily — Excel
# ---------------------------------------------------------------------------

@router.get("/ceo-daily/excel")
async def ceo_daily_excel(
    factory_id: UUID | None = None,
    report_date: str | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_management),
):
    """CEO daily report as a multi-sheet Excel workbook."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    from business.services.daily_kpi import (
        calculate_dashboard_summary,
        calculate_production_metrics,
        calculate_material_metrics,
        get_activity_feed,
    )

    d = date.fromisoformat(report_date) if report_date else date.today()
    summary = calculate_dashboard_summary(db, factory_id, d, d)
    production = calculate_production_metrics(db, factory_id, d, d)
    materials = calculate_material_metrics(db, factory_id)
    activity = get_activity_feed(db, factory_id, limit=30)

    wb = Workbook()
    st = _xl_styles()

    # ---- Sheet 1: KPI Summary ----
    ws = wb.active
    ws.title = "KPI Summary"
    ws.cell(row=1, column=1, value=f"CEO Daily Report — {d}").font = st["section_font"]
    ws.merge_cells("A1:B1")

    kpi_rows = [
        ("Orders In Progress", summary.get("orders_in_progress", 0)),
        ("Total Orders", summary.get("total_orders", 0)),
        ("Output (sqm)", summary.get("output_sqm", 0)),
        ("On-Time Rate (%)", summary.get("on_time_rate", 0)),
        ("Defect Rate (%)", summary.get("defect_rate", 0)),
        ("Kiln Utilization (%)", summary.get("kiln_utilization", 0)),
        ("OEE (%)", summary.get("oee", 0)),
        ("Cost per sqm (USD)", summary.get("cost_per_sqm", 0)),
    ]
    row = 3
    for label, value in kpi_rows:
        ws.cell(row=row, column=1, value=label).font = st["kpi_label_font"]
        ws.cell(row=row, column=1).border = st["thin_border"]
        c = ws.cell(row=row, column=2, value=value)
        c.font = st["kpi_value_font"]
        c.border = st["thin_border"]
        c.alignment = st["center"]
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18

    # ---- Sheet 2: Pipeline Funnel ----
    ws2 = wb.create_sheet("Pipeline")
    row = _write_header_row(ws2, 1, ["Stage", "Positions", "SQM"], st)
    for i, stage in enumerate(production.get("pipeline_funnel", [])):
        row = _write_data_row(ws2, row, [
            stage.get("stage", ""),
            stage.get("count", 0),
            round(stage.get("sqm", 0), 2),
        ], st, alt=(i % 2 == 1))
    _auto_width(ws2)

    # ---- Sheet 3: Critical Positions ----
    ws3 = wb.create_sheet("Critical Positions")
    row = _write_header_row(ws3, 1, [
        "Order #", "Color", "Size", "Qty", "Status", "Delay (hrs)", "Deadline",
    ], st)
    for i, p in enumerate(production.get("critical_positions", [])):
        row = _write_data_row(ws3, row, [
            p.get("order_number", ""),
            p.get("color", ""),
            p.get("size", ""),
            p.get("quantity", 0),
            p.get("status", ""),
            p.get("delay_hours", 0),
            p.get("deadline", ""),
        ], st, alt=(i % 2 == 1))
    _auto_width(ws3)

    # ---- Sheet 4: Material Deficits ----
    ws4 = wb.create_sheet("Material Deficits")
    deficit_items = materials.get("deficit_items", [])
    row = _write_header_row(ws4, 1, [
        "Material", "Type", "Balance", "Min Balance", "Deficit", "Unit",
    ], st)
    for i, m in enumerate(deficit_items):
        row = _write_data_row(ws4, row, [
            m.get("name", ""),
            m.get("material_type", ""),
            m.get("balance", 0),
            m.get("min_balance", 0),
            m.get("deficit", 0),
            m.get("unit", ""),
        ], st, alt=(i % 2 == 1))
    _auto_width(ws4)

    # ---- Sheet 5: Daily Output ----
    daily_output = production.get("daily_output", [])
    if daily_output:
        ws5 = wb.create_sheet("Daily Output")
        row = _write_header_row(ws5, 1, ["Date", "Output (sqm)", "Output (pcs)"], st)
        for i, d_row in enumerate(daily_output):
            row = _write_data_row(ws5, row, [
                d_row.get("date", ""),
                round(d_row.get("output_sqm", 0), 2),
                d_row.get("output_pcs", 0),
            ], st, alt=(i % 2 == 1))
        _auto_width(ws5)

    # ---- Sheet 6: Recent Activity ----
    if activity:
        ws6 = wb.create_sheet("Activity")
        row = _write_header_row(ws6, 1, ["Time", "User", "Action", "Details"], st)
        for i, a in enumerate(activity if isinstance(activity, list) else activity.get("items", [])):
            row = _write_data_row(ws6, row, [
                a.get("timestamp", a.get("created_at", "")),
                a.get("user_name", a.get("user", "")),
                a.get("action", a.get("event_type", "")),
                a.get("details", a.get("description", "")),
            ], st, alt=(i % 2 == 1))
        _auto_width(ws6)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"ceo-daily-{d}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Owner Monthly — Excel
# ---------------------------------------------------------------------------

@router.get("/owner-monthly/excel")
async def owner_monthly_excel(
    factory_id: UUID | None = None,
    month: str | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(require_owner),
):
    """Owner monthly report as a multi-sheet Excel workbook."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    from business.services.daily_kpi import (
        calculate_dashboard_summary,
        calculate_production_metrics,
        calculate_material_metrics,
        calculate_factory_comparison,
    )

    # Period calculation
    if month:
        year, m = int(month[:4]), int(month[5:7])
        d_from = date(year, m, 1)
        if m == 12:
            d_to = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            d_to = date(year, m + 1, 1) - timedelta(days=1)
    else:
        d_to = date.today()
        d_from = d_to.replace(day=1)

    summary = calculate_dashboard_summary(db, factory_id, d_from, d_to)
    production = calculate_production_metrics(db, factory_id, d_from, d_to)
    materials = calculate_material_metrics(db, factory_id)

    # Financial breakdown
    financial_q = db.query(
        FinancialEntry.entry_type,
        FinancialEntry.category,
        sa_func.sum(FinancialEntry.amount).label("total"),
    ).filter(
        FinancialEntry.entry_date >= d_from,
        FinancialEntry.entry_date <= d_to,
    )
    if factory_id:
        financial_q = financial_q.filter(FinancialEntry.factory_id == factory_id)
    financial_q = financial_q.group_by(FinancialEntry.entry_type, FinancialEntry.category)

    financials_flat = []
    for row in financial_q.all():
        entry_type = row.entry_type if isinstance(row.entry_type, str) else row.entry_type.value
        category = row.category if isinstance(row.category, str) else row.category.value
        financials_flat.append({
            "type": entry_type,
            "category": category,
            "total": float(row.total or 0),
        })

    # Revenue
    revenue_q = db.query(
        sa_func.sum(OrderFinancial.total_price)
    ).join(ProductionOrder).filter(
        ProductionOrder.shipped_at.isnot(None),
        sa_func.date(ProductionOrder.shipped_at) >= d_from,
        sa_func.date(ProductionOrder.shipped_at) <= d_to,
    )
    if factory_id:
        revenue_q = revenue_q.filter(ProductionOrder.factory_id == factory_id)
    revenue = float(revenue_q.scalar() or 0)

    # Factory comparison
    factory_comparison = calculate_factory_comparison(db)

    wb = Workbook()
    st = _xl_styles()

    # ---- Sheet 1: Monthly Summary ----
    ws = wb.active
    ws.title = "Monthly Summary"
    period_label = f"{d_from.strftime('%B %Y')}" if month else f"{d_from} — {d_to}"
    ws.cell(row=1, column=1, value=f"Owner Monthly Report — {period_label}").font = st["section_font"]
    ws.merge_cells("A1:B1")

    kpi_rows = [
        ("Period", f"{d_from} to {d_to}"),
        ("Revenue (USD)", f"${revenue:,.2f}"),
        ("Orders In Progress", summary.get("orders_in_progress", 0)),
        ("Total Orders", summary.get("total_orders", 0)),
        ("Output (sqm)", summary.get("output_sqm", 0)),
        ("On-Time Rate (%)", summary.get("on_time_rate", 0)),
        ("Defect Rate (%)", summary.get("defect_rate", 0)),
        ("Kiln Utilization (%)", summary.get("kiln_utilization", 0)),
        ("OEE (%)", summary.get("oee", 0)),
        ("Cost per sqm (USD)", summary.get("cost_per_sqm", 0)),
    ]
    row = 3
    for label, value in kpi_rows:
        ws.cell(row=row, column=1, value=label).font = st["kpi_label_font"]
        ws.cell(row=row, column=1).border = st["thin_border"]
        c = ws.cell(row=row, column=2, value=value)
        c.font = st["kpi_value_font"]
        c.border = st["thin_border"]
        c.alignment = st["center"]
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 22

    # ---- Sheet 2: Financial Breakdown ----
    ws2 = wb.create_sheet("Financials")
    ws2.cell(row=1, column=1, value=f"Revenue: ${revenue:,.2f}").font = st["section_font"]
    ws2.merge_cells("A1:C1")
    row = 3
    row = _write_header_row(ws2, row, ["Type", "Category", "Amount (USD)"], st)
    total_expenses = 0
    for i, f_row in enumerate(sorted(financials_flat, key=lambda x: (x["type"], x["category"]))):
        row = _write_data_row(ws2, row, [
            f_row["type"].upper(),
            f_row["category"],
            round(f_row["total"], 2),
        ], st, alt=(i % 2 == 1))
        total_expenses += f_row["total"]

    # Totals row
    from openpyxl.styles import Font
    ws2.cell(row=row, column=1, value="").border = st["thin_border"]
    ws2.cell(row=row, column=2, value="TOTAL EXPENSES").font = Font(bold=True, size=11)
    ws2.cell(row=row, column=2).border = st["thin_border"]
    tc = ws2.cell(row=row, column=3, value=round(total_expenses, 2))
    tc.font = Font(bold=True, size=11)
    tc.border = st["thin_border"]
    row += 1
    ws2.cell(row=row, column=2, value="NET PROFIT").font = Font(bold=True, color="2B5797", size=11)
    ws2.cell(row=row, column=2).border = st["thin_border"]
    pc = ws2.cell(row=row, column=3, value=round(revenue - total_expenses, 2))
    pc.font = Font(bold=True, color="2B5797", size=11)
    pc.border = st["thin_border"]
    _auto_width(ws2)

    # ---- Sheet 3: Factory Comparison ----
    if factory_comparison:
        ws3 = wb.create_sheet("Factory Comparison")
        row = _write_header_row(ws3, 1, [
            "Factory", "Location", "Orders", "Output (sqm)",
            "On-Time %", "Defect %", "Kiln Util %", "OEE %", "Cost/sqm",
        ], st)
        for i, fc in enumerate(factory_comparison):
            row = _write_data_row(ws3, row, [
                fc.get("factory_name", ""),
                fc.get("factory_location", ""),
                fc.get("orders_in_progress", 0),
                fc.get("output_sqm", 0),
                fc.get("on_time_rate", 0),
                fc.get("defect_rate", 0),
                fc.get("kiln_utilization", 0),
                fc.get("oee", 0),
                fc.get("cost_per_sqm", 0),
            ], st, alt=(i % 2 == 1))
        _auto_width(ws3)

    # ---- Sheet 4: Production Pipeline ----
    ws4 = wb.create_sheet("Pipeline")
    row = _write_header_row(ws4, 1, ["Stage", "Positions", "SQM"], st)
    for i, stage in enumerate(production.get("pipeline_funnel", [])):
        row = _write_data_row(ws4, row, [
            stage.get("stage", ""),
            stage.get("count", 0),
            round(stage.get("sqm", 0), 2),
        ], st, alt=(i % 2 == 1))
    _auto_width(ws4)

    # ---- Sheet 5: Critical Positions ----
    ws5 = wb.create_sheet("Critical Positions")
    row = _write_header_row(ws5, 1, [
        "Order #", "Color", "Size", "Qty", "Status", "Delay (hrs)", "Deadline",
    ], st)
    for i, p in enumerate(production.get("critical_positions", [])):
        row = _write_data_row(ws5, row, [
            p.get("order_number", ""),
            p.get("color", ""),
            p.get("size", ""),
            p.get("quantity", 0),
            p.get("status", ""),
            p.get("delay_hours", 0),
            p.get("deadline", ""),
        ], st, alt=(i % 2 == 1))
    _auto_width(ws5)

    # ---- Sheet 6: Material Deficits ----
    deficit_items = materials.get("deficit_items", [])
    if deficit_items:
        ws6 = wb.create_sheet("Material Deficits")
        row = _write_header_row(ws6, 1, [
            "Material", "Type", "Balance", "Min Balance", "Deficit", "Unit",
        ], st)
        for i, m in enumerate(deficit_items):
            row = _write_data_row(ws6, row, [
                m.get("name", ""),
                m.get("material_type", ""),
                m.get("balance", 0),
                m.get("min_balance", 0),
                m.get("deficit", 0),
                m.get("unit", ""),
            ], st, alt=(i % 2 == 1))
        _auto_width(ws6)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    month_label = month or d_from.strftime("%Y-%m")
    filename = f"owner-monthly-{month_label}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
